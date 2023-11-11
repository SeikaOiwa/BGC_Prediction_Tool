import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import glob
import os
import shutil
from io import BytesIO
import subprocess
import time
import numpy as np

# docking simulaton
import oddt
from oddt import toolkit
from oddt import docking
from rdkit import Chem
from rdkit.Chem import AllChem, Draw, Descriptors, PandasTools
from vina import Vina

# pubchem検索
import pubchempy as pcp
from PIL import Image

# pubmed検索
from Bio import Entrez
Entrez.email = "seika.ooiwa@gmail.com"

#path情報
base_path = os.getcwd()
i_path = f'{base_path}/images'
p_path = f'{base_path}/scripts'
fig_save_path = f'{base_path}/static'
enzyme_path = f'{base_path}/Docking/enzyme'
ligand_path = f'{base_path}/Docking/ligand'
opath = f'{base_path}/Docking'
fpath = f'{base_path}/Genome'
log_path = f'{base_path}/log'
download_path = base_path

#ゲノム情報取得

def get_genome(species,group,base_path):
    """NCBIにアクセスし、Bacterial nameからゲノムデータ（**.gbff.gz)を検索・ダウンロード
       NCBI上に登録が無く空フォルダが生成した場合は削除
    　 ゲノムデータの登録有無を返す
    
    Parameters:
    ----------
    species: str
       生物名 (ex. Bacillus subtilis)
    group: str
       bacteria or fungi or plant name 
    base_path: str
       Antismashフォルダパス (ex. /home/***/notebooks/Mydata_analysis/Antismash)
       
    Returns:
    --------
    registration: str
       ゲノム情報登録の有無
    """
    # データ格納用フォルダ作成
    sfname = species.replace(' ','_')
    genome_fpath = f'{base_path}/Genome/{sfname}'
    
    # NCBIからゲノム情報を取得
    target = '"'+species+'"'+' '+group
    subprocess.run(f'ncbi-genome-download -g {target} --refseq-categories representative,reference -F genbank -o tmp',shell=True)
    
    # NCBI登録有無
    registration = '無'

    # ダウンロードしたゲノム情報（.gzファイル）を格納用フォルダに移動、残りは削除
    if os.path.isdir(f'{base_path}/tmp'):
        fpath = glob.glob(f'{base_path}/tmp/**/*.gz',recursive=True)[0]
        if os.path.isfile(fpath):
            os.makedirs(f'{genome_fpath}',exist_ok=True)
            registration = '有'
            shutil.move(f'{fpath}',f'{genome_fpath}')
            shutil.rmtree(f'{base_path}/tmp')
        else:
            shutil.rmtree(f'{base_path}/tmp')

    return registration

# Anti_smash

def omix_data(result_path):
    """Antismash.csvデータをconcat
       result_path下の各フォルダからantismash.csvを読込みconcatし、omix_dataを生成

    Parameters:
    -----------
    result_path: str 
       antismash解析データ保存フォルダのパス
          
    Returns: 
    --------
    omix_data: DataFrame
    """
    dir_list = os.listdir(f'{result_path}')
    
    result_list = []
    
    for sfile in dir_list:
        data_path = f'{result_path}/{sfile}/antismash.csv'
    
        if os.path.isfile(data_path) :
            df = pd.read_csv(data_path)
            result_list.append(df)
    
    omix_data = pd.concat(result_list) 
    
    return omix_data
    
def create_structure(smiles,target,fig_save_path):
    """smilesから2D構造データ(.png)を作成
       Streamlitで表示するためstaticフォルダ下に保存

    Paramerters
    -----------
    smiles: str
        smiles表記のデータ 
    target: str
        smilesに該当する物質名
    fig_save_path: str
        データ保存先：(ex. /home/***/notebooks/Mydata_analysis/static/antismash)

    See Also
    --------
    add_structure: Antismashで推測した二次代謝物リストからsmilesデータを取得
                   インプットのDataFrameにsmilesを追加
    """

    mol = Chem.MolFromSmiles(smiles)
    Draw.MolToFile(mol,f'{fig_save_path}/{target}.png',size=(100,100))


def make_output_data(omix_data,base_path,save_name):
    """milesから2D構造を作成、dataframeに埋込み、'Bacterial_list.xlsx'で保存

    Prameters:
    ----------
    omix_data: DataFrame
        smiles含むデータフレーム *列名はsmilesとする
    base_path: str
        解析データの書出し先のパス(ex. /home/***/notebooks/Mydata_analysis/Antismash)
    save_name: str
        ファイル保存名(拡張子は含まない)

    Returns:
    -------
    omix_data_withfig: DataFrame
        2D構造を埋め込んだデータフレーム（jupyter notebook上で画像が表示される）
    """
    
    # NaNをスペースで埋める
    omix_data_withfig = omix_data.fillna('')
    # smilesから構造データ作成
    omix_data['Structure']=''
    PandasTools.AddMoleculeColumnToFrame(omix_data_withfig,molCol='Structure',smilesCol='smiles')
    # excel形式で保存
    PandasTools.SaveXlsxFromFrame(omix_data_withfig,f'{base_path}/{save_name}.xlsx',molCol='Structure',size=(150,150))
    
    return omix_data_withfig

def get_smile_data(target):
    """PubChemにアクセスし、対象物質(target)に該当するCanonicalSMILESデータを取得
       見つからない場合はNoneを返す
    
    Parameters:
    ----------
    target: str
        smiles検索する化合物名

    Returns:
    -------
    cid: str
       chemical_id
    smiles: str
       smiles表記

    See Also:
    --------
    add_structure: Antismashで推測した二次代謝物リストからsmilesデータを取得
    """
    
    properties = ['CanonicalSMILES']
    
    # pubChemからデータ読み込み
    c = pcp.get_properties(properties, f'{target}','name')
    if not c:
        cid=None
        smiles=None
    else:
        df = pd.DataFrame(c)
        # CID,smilesデータ抽出
        cid = df.loc[0,'CID']
        smiles = df.loc[0,'CanonicalSMILES']
        
    return cid,smiles

def add_structure(antismash_result):    
    """Antismashで推測した二次代謝物リストからsmilesデータを取得
       インプットのDataFrameにsmilesを追加
    
    Prameters:
    ---------
    antismash_result: DataFrame
        'Candidate_metabolite'列に物質名を含むデータフレーム

    Returns:  
    --------
    merge_data2: DataFrame

    See Also:
    --------
    get_smile_data: PubChemにアクセスし、対象物質(target)に該当するCanonicalSMILESデータを取得
    create_structure: smilesから2D構造データ(.png)を作成
    """
    
    smile_list = []

    for i in range(len(antismash_result)):
        target = antismash_result.loc[i,'Candidate_metabolite']
        
        # pubChemからsmilesデータ取得
        cid,smiles = get_smile_data(target)

        if cid is not None:
            # smilesデータから2D構造作成(streamlit表示用)
            fig_path = f'./app/static/{target}.png'
            create_structure(smiles,target,fig_save_path)
            # リスト追加
            add = [target,cid,smiles,fig_path]
            smile_list.append(add)

    # smileデータフレーム作成
    add_data = pd.DataFrame(smile_list,columns=['target','CID','smiles','apps'])

    # antismashデータにsmiles追加
    merge_data = pd.merge(antismash_result,add_data,left_on='Candidate_metabolite',right_on='target',how='left')
    merge_data2 = merge_data[['Bacterial_name','Type','Candidate_metabolite','Similarity','CID','smiles','apps']] # appsはstreamlit表示用
    
    return merge_data2

def create_query():
    """pubmed検索用のクエリー作成に用いるデータセット（機能名：定義）を生成
    
    Returns:
    --------
    category: list
        機能名のリスト
        
    func_name_to_query: dict
        機能名と定義（辞書型）ex. {"anti_bacteria/fungi":['antibacterial','antimicrobial','antifungal','antimycotic','fungicide'],,}
    """
    
    # 機能の定義
    func_name_to_keywords = {
        'anti_bacteria/fungi':['antibacterial','antimicrobial','antifungal','antimycotic','fungicide'],
        'insecticide':['insecticidal','insecticide','pesticide'],
        'herbicide':['herbicide','weed killer','weedkiller'],
        'tickkiller':['acaricidal','tick killer','tickkiller'],
        'virus':['virucidal','antivirus','antiviral'],
    }
    # queryの生成
    func_name_to_query = {
        func_name: get_query(keywords) for func_name, keywords in func_name_to_keywords.items()
    }

    category = list(func_name_to_query.keys())

    return category,func_name_to_query
    

def ex_antismash(result_anti_path,bacterial_name):

    """Antismash解析データ群から、既知二次代謝物名のリストを作成

    Parameterts:
    -----------
    result_anti_path : str
        Antismash解析データ(index.html)のフォルダパス
    bacterial_name : str
        リスト記載の生物名, データフレームに追加用

    Reterns:
    -------
    antismash_result: DataFrame
        bacterial_name,types,active_agent,similarityが記載
    total_cluster_num: int
        Antismashで検出されたクラスター総数
    knw_num: int
        既知二次代謝物に相同性が認められたクラスター数
    unknw_num: int
        未同定のクラスター数
    """

    #tmpフォルダ内のhtmlを読み込み
    if os.path.isfile(f'{result_anti_path}/index.html'):
        result = pd.read_html(f'{result_anti_path}/index.html')

        result_list = []
        total_cluster_num = 0
        knw_num = 0

        #データ抽出 
        for i in range(0,len(result)):
            data = result[i]
            if 'Most similar known cluster' in data.columns:
                # 全クラスター数のカウント
                total_cluster_num = total_cluster_num + len(data)
                
                #既知物質と類似性が認められたものを抽出
                data2 = data.dropna(subset=['Most similar known cluster']).reset_index()
                # クラスター数のカウント
                knw_num = knw_num + len(data2)
                
                for i2 in range(len(data2)):
                    #必要な要素を抽出
                    types = data2.loc[i2,'Type']
                    active_agent = data2.loc[i2,'Most similar known cluster']
                    similarity = int(data2.loc[i2,'Similarity'].rstrip('%')) #int(data2.loc[i2,'Similarity'][:-1])
            
                    #リストに追加
                    add_data = [bacterial_name,types,active_agent,similarity]
                    result_list.append(add_data)
            
        #データフレーム作成
        antismash_result_ = pd.DataFrame(result_list,columns=['Bacterial_name','Type','Candidate_metabolite','Similarity'])
       
        # 同じ物質名は削除
        antismash_result = antismash_result_.drop_duplicates(subset=['Candidate_metabolite']).reset_index()

        #tmpフォルダ削除 (同名フォルダが存在するとantismashでエラーが出る)
        shutil.rmtree(result_anti_path)
    else:
        shutil.rmtree(result_anti_path)
    
    unknw_num = total_cluster_num - knw_num
    
    return antismash_result,total_cluster_num,knw_num, unknw_num


def ex_docking_data(rpath):
    """Docking SimulationデータからdG値を抽出

    Parameters:
    ----------
    rpath : str
        Docking Simulationデータパス (ex. /tmp.pdbqt)

    Returns:
    -------
    min_dg : str

    See Also:
    --------
    add_docking_data
    """
    df = pd.read_table(rpath)
    df2 = df[df['MODEL 1'].str.contains('RESULT')]
    df3 = df2.reset_index(drop=True)
    df4 = df3['MODEL 1'].str.split(':',expand=True)
    df5 = df4[1].str.split('      ',expand=True)
    min_dg = df5.loc[0,0].replace(' ','')

    return min_dg


def create_3d_structure(p_path,smile_data,ligand_path):
    """dataframe内のsmilesからdocking用の3D構造データを作成し、保存（{ligand_path}/ligand.pdbqt)
       smile -> mol -> 3D .sdf -> add H2,charge -> ligand.pdbqt

    Parameters:
    ----------
    p_path: str
        convert_to_pdpqt.pyファイルの保存先パス
    smile_data : str
        smileデータ
    ligand_path : str
        リガンドデータ保存先(ex. /Antismash/Docking/ligand)

    See Also:
    --------
    add_docking_data : Antismash解析データから構築した既知二次代謝物をDocking simulation、結果を追加したデータフレームを返す
    """
    
    subprocess.run(["python",f"{p_path}/convert_to_sdf.py",smile_data,ligand_path])
    subprocess.run(["python",f"{p_path}/convert_to_pdbqt.py",ligand_path])
    
def docking_simulation(denzyme_path,dligand_path,binding_cite,search_area,exhaustiveness,data_num,opath):
    """autdock_vina を使ったdocking simulationを行い、dGを含むtmp.pdbqtが生成
       1回のドッキングシミュレーションの制限時間は20分、それ以上かかる場合は結合しないものと判断する

    Parameters:
    ----------
    denzyme_path: str
      enzyme data　path (/**.pdbqt)
    dligand_path: str
      ligand data　path (/**.pdbqt)
    binding_cite: list,float
      基質結合位置情報, [x,y,z]
    search_area: list,float
      リガンドの結合の探索範囲 [x,y,z]
    exhaustiveness: int
      探索回数, 低いと予測精度が落ちる（50>に設定)
    data_num: int
      取得データ数
    opath: str
      simulation結果の保存フォルダのパス

    See Also:
    --------
    add_docking_data : Antismash解析データから構築した既知二次代謝物をDocking simulation、結果を追加したデータフレームを返す
    """
    
    try:
        subprocess.run(["python",f"{p_path}/docking.py",denzyme_path,
                dligand_path,
                ",".join(str(num) for num in binding_cite),
                ",".join(str(num2) for num2 in search_area),
                str(exhaustiveness),str(data_num),opath],timeout=int(search_time))
        return 'finish'
    except:
    #except subprocess.TimeoutExpired as e:
        return 'not_docking'
        
def get_docking_condition(path) :
    """酵素毎に設定したドッキング条件を抽出
      ドッキング条件はDocking_condition.csvに記載
      - 基質結合部位の中心座標（予測値）[ProteinsPlus](https://proteins.plus/)
      -> binding_cite = [lx,ly,lz]
      - 探索範囲（x,y,z）
      -> search_area = [sx,sy,sz]
      - 探索回数(exhaustiveness) ※低いと予測精度が落ちる（50>に設定)
      -> ex_num
      - 取得データ数 10コ程度
      -> data_num
      - 基準となるdG値, 予めdocking simulationしておいたオリジナルのligandのdG値
      -> original_dg

    Parameters:
    ----------
    path: str
       ドッキング条件データ(docking_condition.csv) のパス

    Returns:
    -------
    binding_cite : list,float
    search_area : list, float
    exhaustiveness : int
    data_num : int
    original_dg : int

    See Also:
    --------
    add_docking_data : Antismash解析データから構築した既知二次代謝物をDocking simulation、結果を追加したデータフレームを返す
    """
    df = pd.read_csv(path,skiprows=1)
    lx = float(df.loc[0,'lx'])
    ly = float(df.loc[0,'ly'])
    lz = float(df.loc[0,'lz'])
    binding_cite = [lx,ly,lz]
    sx = float(df.loc[0,'sx'])
    sy = float(df.loc[0,'sy'])
    sz = float(df.loc[0,'sz'])
    search_area = [sx,sy,sz]
    exhaustiveness = int(df.loc[0,'ex_num'])
    data_num = int(df.loc[0,'data_num'])
    original_dg = int(df.loc[0,'original_dg'])
    
    return binding_cite, search_area, exhaustiveness, data_num, original_dg

def add_docking_data(asdf,enzyme_path,enzyme_list_name,ligand_path,opath,search_time):

    """AutoDock Vina によるDocking Simulationを実施
    オリジナル基質のdelta Gとの比較から、ligandの結合力を３段階(High,Middle,Low)で評価
    Input DataFrameに評価結果を追加

    Parameters:
    ----------
    asdf : DataFrame
        Antismash解析データフレーム,  列名"Candidate_metabolite" : ligand name, 列名"smiles" : smiles data
    enzyme_path : str
        酵素データフォルダのパス(ex. /Antismash/Docking/enzyme)
    enzyme_list_name : list
        酵素データ名のリスト, enzymeフォルダ下にリスト記載名に一致するDirectoryが存在すること
        Directory下に酵素データ(*.pdbqt)、Docking condition(docking_condition.csv)が格納
    ligand_path : str
        リガンドデータ保存先(ex. /Antismash/Docking/ligand)
    opath: str
        docking simulationデータの保存フォルダのパス
    search_time: int/str
      Docking simulation (subprocess.run) の実行時間の上限値
      
    Returns:
    -------
    asdf: DataFrame

    See Also:
    --------
    create_3d_structure
    get_docking_condition
    docking_simulation
    ex_docking_data
    """
    # asdfのNanを'none'で埋める
    asdf2 = asdf.fillna('none')
    # asdfにdocking simulation結果入力カラムを追加
    asdf[enzyme_list_name] = ''
    
    for i in range(len(asdf2)):
        # 要素抽出
        smile_data = asdf2.loc[i,'smiles']
        ligand = asdf2.loc[i,'Candidate_metabolite']
        
        # smilesデータ記載されていれば、3D構造構築
        if smile_data != 'none':
            create_3d_structure(p_path,smile_data,ligand_path)
            
            # ligand.pdpqtが生成されれば、以下を実行
            if os.path.isfile(f'{ligand_path}/ligand.pdbqt'):
                
                for en in enzyme_list_name:
                    path = f'{enzyme_path}/{en}/docking_condition.csv'
                    denzyme_path = f'{enzyme_path}/{en}/{en}.pdbqt'
                    dligand_path = f'{ligand_path}/ligand.pdbqt'

                    #酵素毎に設定したドッキング条件を呼出し
                    binding_cite, search_area, exhaustiveness, data_num, original_dg = get_docking_condition(path)  
                    #ドッキングデータ取得
                    search_time = 900
                    result_docking = docking_simulation(denzyme_path,dligand_path,binding_cite,search_area,exhaustiveness,data_num,opath,search_time)

                    if result_docking == 'finish':
                        if os.path.isfile(f'{opath}/tmp.pdbqt'):
                            #結果抽出
                            min_dg = ex_docking_data(f'{opath}/tmp.pdbqt')
                            #基準dG(オリジナル基質)との比較
                            diff = -(float(min_dg) - float(original_dg))
                            if 0.5 < diff:
                                input_dg = 'High'
                            elif -0.5 <= diff <= 0.5:
                                input_dg = 'Middle'
                            elif diff < -0.5:
                                input_dg = 'Low'

                            #データフレームへの書き込み
                            asdf.loc[i,en]=input_dg

                            #不要ファイルの削除
                            os.remove(f'{ligand_path}/ligand.pdbqt')
                            os.remove(f'{opath}/tmp.pdbqt')
                        else:
                            #データフレームへの書き込み
                            asdf.loc[i,en]='not docking'
                            os.remove(f'{ligand_path}/ligand.pdbqt')
                            
                    if result_docking == 'not_docking':
                        #データフレームへの書き込み
                        asdf.loc[i,en]='not docking'
                        os.remove(f'{ligand_path}/ligand.pdbqt')
            else:
                pass

    return asdf
   
def gain_genome():
   st.markdown("""
   ## STEP1: ゲノム情報の取得  
   #### 下記の要素を有する生物名リストをアップロード
   - species: 生物名（ex. Bacillus subtilis)
   - group: グループ（bacteria,fungi,plantから選択）
   """)
   
   # uploader
   bacterial_name_list = st.file_uploader('',type='csv',key='f')
   if bacterial_name_list:
      bac_name_list = pd.read_csv(bacterial_name_list)
      st.dataframe(bac_name_list)
   
   # 生物名、グループを抽出、ゲノム情報を取得
   st.markdown("""ゲノム情報の取得""")
   go = st.button('情報取得')
   if go:  
      bac_name_list['get_result'] = ''
      # log用
      bac_name_list.to_csv(f'{log_path}/genome_log.csv')
      bac_name_list['progress(sec)'] = ''
     
      for i in range(len(bac_name_list)):
         start = time.time()
         species = bac_name_list.loc[i,'species']
         group = bac_name_list.loc[i,'group']
         registration = get_genome(species,group,base_path)
         
         #ゲノム情報取得の有無をリストに追加
         bac_name_list.loc[i,'get_result'] = registration
         end = time.time()
         # logに進行状況を記録
         log = pd.read_csv(f'{log_path}/genome_log.csv')
         log.loc[i,'progress'] = round(end - start,2)
         log.to_csv(f'{log_path}/genome_log.csv')

      st.dataframe(bac_name_list)
      
      # 菌株リスト+ゲノム情報取得結果を保存
      bac_name_list.to_csv(f'{fpath}/download_genome_list.csv')
      # downloader
      bac_name_list.to_excel(buf := BytesIO(), index=True)
      st.download_button("Download",buf.getvalue(),f"download_genome_list","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def get_query(keywords,field="[Title/Abstract]"):
    """keywords中の機能名リストを取り出し、field名と連結した上で、機能名同しをOR連結する
    
    Parameters
    ----------
    keywords: list
        機能名のリスト
    field: str
        Pubmed上の検索領域（default: Title/Abstract)
    Returns
    -------
    "(" + ors + ")": str
        検索文字列
    """
    ors = " OR ".join(f"({keyword}{field})" for keyword in keywords)
    return "(" + ors + ")"

def create_query():
    """pubmed検索用のクエリー作成に用いるデータセット（機能名：定義）を生成
    
    Returns:
    --------
    category: list
        機能名のリスト
        
    func_name_to_query: dict
        機能名と定義（辞書型）ex. {"anti_bacteria/fungi":['antibacterial','antimicrobial','antifungal','antimycotic','fungicide'],,}
    """
    
    # 機能の定義
    func_name_to_keywords = {
        'anti_bacteria/fungi':['antibacterial','antimicrobial','antifungal','antimycotic','fungicide'],
        'insecticide':['insecticidal','insecticide','pesticide'],
        'herbicide':['herbicide','weed killer','weedkiller'],
        'tickkiller':['acaricidal','tick killer','tickkiller'],
        'virus':['virucidal','antivirus','antiviral'],
    }
    # queryの生成
    func_name_to_query = {
        func_name: get_query(keywords) for func_name, keywords in func_name_to_keywords.items()
    }

    category = list(func_name_to_query.keys())

    return category,func_name_to_query

def ex_pubmed(asdf,col_name,fname_list,sign):
    """データフレーム中のcol_nameの項目をグループ化し、各要素の数をカウント
    　　カウントする要素は１種類（signで指定）
    　　col_nameとfname_listの項目を含むデータフレームを返す

    Parameters
    ----------
    asdf: dataframe
    col_name: str
        groupbyメソッドを適用するカラム名
    fname_list: list
        asdfから抽出するカラム名
    sign: str
        カウント対象となる要素の表記内容
    Returns
    -------
    ex_p2: dataframe
    """
    data = []
    for i in fname_list:
        exdata = asdf[i].apply(lambda x: 1 if x==sign else 0)
        data.append(exdata)

    tmp_df = pd.DataFrame(data)
    tmp_df2 = tmp_df.T
    tmp_df2[col_name] = asdf[col_name]

    ex_p = tmp_df2.groupby(col_name).sum().reset_index()

    return ex_p

def adjust_xlsx(path_to_excel,sheet_n,c_dim):
    """与えられたエクセルの書式を変更
    　　・縦方向（中央揃え）
    　　・カラム幅（c_dimで指定）
    Prameter:
    --------
    path_to_excel: str
        出力済みのxlsxファイルへのパス（'/usr/virtual_screening/***.xlsx')
    sheet_n: str
        シート名
    c_dim: dict
        {列名(ex.'A'): 幅}
    
    """
    # データ読み込み
    wb = load_workbook(f'{path_to_excel}')
    ws = wb[sheet_n]
    
    # 縦方向；中央揃え
    ws.alignment = Alignment(vertical="center")

    # 折り返して全体表示
    ws.alignment = Alignment(wrapText=True)    

    # カラム横幅の調整
    for colm,wnum in c_dim.items():
        ws.column_dimentions[colm].width = wnum

    # 保存 
    wb.save(f'{path_to_excel}')

def add_df_to_excel(path_to_excel,add_dataframe,sheet_n):
    """ 出力済みのxlsxファイルの別シートに与えられたデータフレームから全データを追加
    
    Prameters:
    ---------
    path_to_excel: str
        出力済みのxlsxファイルへのパス（'/usr/virtual_screening/***.xlsx')
    add_dataframe: dataframe
        追加するデータフレーム
    sheet_n: str
        追加するシート名
    """
    with pd.ExcelWriter(f'{path_to_excel}',mode='a',engine='openpyxl') as writer:
        add_dataframe.to_excel(writer, sheet_name=sheet_n,index=False)
    # データフレーム追加したシートを左に移動（＊outlineを対象とするため）
    wb = load_workbook(f'{path_to_excel}')
    wb.move_sheet(sheet_n, offset=-1)
    wb.save(f'{path_to_excel}')

def anti_smash():      
   st.markdown(""" ####
   - 「Genome」フォルダに対象のゲノムデータを保存（STEP1を実施した場合は不要）""")    
  
   # ゲノムリストの抽出（フォルダ名＝ゲノム名になっている）
   dir_list = [f for f in os.listdir(fpath) if os.path.isdir(os.path.join(fpath,f))]
   
   # outline用のdataframe
   # outline = pd.DataFrame(dir_list,columns=['Bacterial_name']).set_index('Bacterial_name')
   # outline[['known_cluster','unknown_cluster','total_cluster']]=''

   # log記録用データフレーム
   log = pd.DataFrame(data = dir_list,columns=['Bacterial_name']).set_index('Bacterial_name')
   log['progress(min)'] = ''
   log['data_num'] = ''
   log.to_csv(f'{log_path}/antismash_log.csv')
   
   go2 = st.button('解析開始')
   if go2:
        
      # total_time
      t_start = time.time()

      for genome in dir_list:
         genome_path = glob.glob(f'{fpath}/{genome}/*') 

         # log用
         start = time.time()
         # antismash実行
         subprocess.run(f'antismash {genome_path[0]} --cb-knownclusters --output-dir {base_path}/tmp_anti',shell=True)

         # antismash解析データ抽出
         antismash_result,total_cluster_num,knw_num,unknw_num = ex_antismash(f'{base_path}/tmp_anti',genome)

         # outlineにクラスター情報、機能推定結果を追記
         outline = pd.DataFrame(data=[[genome,'','','']],
         columns=['Bacterial_name','known_cluster','unknown_cluster','total_cluster']).set_index('Bacterial_name')
         outline.loc[genome,'known_cluster']=knw_num
         outline.loc[genome,'unknown_cluster']=unknw_num
         outline.loc[genome,'total_cluster']=total_cluster_num

         # smilesデータ追加        
         merge_data2 = add_structure(antismash_result)

         # データ保存
         os.makedirs(f'{fpath}/{genome}/Antismash',exist_ok=True)
         merge_data2.to_csv(f'{fpath}/{genome}/Antismash/antismash.csv',index=False)
         outline.to_csv(f'{fpath}/{genome}/Antismash/outline.csv')

         # 進捗を記録
         end = time.time()
         log = pd.read_csv(f'{log_path}/antismash_log.csv').set_index('Bacterial_name')
         log.loc[genome,'progress(min)'] = round((end - start)/60,2)
         log.loc[genome,'data_num'] = total_cluster_num
         log.to_csv(f'{log_path}/antismash_log.csv')
        
      t_end = time.time()
      total_time = round((t_end - t_start)/60,1)
      st.write(f'解析時間(分)_{total_time}')
    
def add_research_article():
   st.markdown(""" 
   ##### 対象：「Antismash」フォルダ内の”antismash.csvに記載の二次代謝物名
   """) 

   # ゲノムリストの抽出（フォルダ名＝ゲノム名になっている）
   dir_list = [f for f in os.listdir(fpath) if os.path.isdir(os.path.join(fpath,f))]

   # log記録用データフレーム
   log = pd.DataFrame(data = dir_list,columns=['Bacterial_name'])
   log['progress(min)'] = ''
   log['data_num'] = ''
   log.to_csv(f'{log_path}/pred_log.csv')

   go3 = st.button('機能推定スタート')
   if go3:
      # total time
      t_start = time.time()
      for genome in dir_list:
         antidata_Fpath = f'{fpath}/{genome}' 

         # log用
         start = time.time()

         if antidata_Fpath:
            # 結果保存用フォルダ(Pred_func)作成
            os.makedirs(f'{fpath}/{genome}/Pred_func',exist_ok=True)
            # pubmed検索
            subprocess.run(["python",f"{p_path}/pubmed.py",f'{antidata_Fpath}/Antismash/antismash.csv',f'{antidata_Fpath}/Pred_func/antismash.csv'])

            # pubmed検索結果の概要抽出とoutlineへの入力
            asdf = pd.read_csv(f'{antidata_Fpath}/Pred_func/antismash.csv')
            outline = pd.read_csv(f'{antidata_Fpath}/Antismash/outline.csv')
            category,func_name_to_query = create_query()
            ex_p = ex_pubmed(asdf,col_name='Bacterial_name',fname_list=category,sign='Hit')
            outline_p = pd.merge(outline,ex_p,on = 'Bacterial_name',how='left')

            # データ保存
            outline_p.to_csv(f'{fpath}/{genome}/Pred_func/outline.csv',index=False)

            # 進捗を記録
            end = time.time()
            log = pd.read_csv(f'{log_path}/pred_log.csv').set_index('Bacterial_name')
            log.loc[genome,'progress(min)'] = round((end - start)/60,2)
            log.loc[genome,'data_num'] = len(asdf)
            log.to_csv(f'{log_path}/pred_log.csv')
      t_end = time.time()
      total_time = round((t_end - t_start)/60,1)
      st.write(f'解析時間(分)_{total_time}')

def docking():
    st.markdown(""" ## STEP3: Docking simulation  
   #### Step2で生成したExcelデータをアップロード（下記要素を含めば良い）
   - Candidate_metabolite: 予測された二次代謝物名
   - smiles: smiles *smilesデータが存在する二次代謝物のみdockingを実行
   """)
   
   # uploader
    input_data = st.file_uploader('',type='xlsx',key='g')
    if input_data:
       docking_list = pd.read_excel(input_data)
       st.dataframe(docking_list)
    
    st.markdown(""" ##### dockingに用いる酵素を選択
    """)
    e_list = [f for f in os.listdir(enzyme_path) if os.path.isdir(os.path.join(enzyme_path,f))]
    select_en = st.multiselect('',e_list)   

    go5 = st.button('dockin開始')
    if go5:
        print('')
    
#add_docking_data(asdf,enzyme_path,enzyme_list_name,ligand_path,opath,search_time)

def create_final_data(target_process):
   st.markdown(""" 
   二次代謝遺伝子の予測〜Docking結果を統合し、smilesから生成した2D構造を追加したExcelを作成
   """)        
   st.markdown(""" ####
    - 保存ファイル名を入力""")
   sv_name = st.text_input('','antismash_analysis',key = 'a') # default: antismash_analysis

   # ゲノムリストの抽出（フォルダ名＝ゲノム名になっている）
   dir_list = [f for f in os.listdir(fpath) if os.path.isdir(os.path.join(fpath,f))]

   go4 = st.button('開始')
   if go4:

      data_all = []
      outline_all = []

      for genome in dir_list:
        # path作成
        anti_data_path = f'{fpath}/{genome}/{target_process}/antismash.csv'
        outline_path = f'{fpath}/{genome}/{target_process}/outline.csv'

        # データ統合
        if anti_data_path:
            tmp_df = pd.read_csv(anti_data_path)
            outline = pd.read_csv(outline_path)

            data_all.append(tmp_df)
            outline_all.append(outline)     
        else:
            pass
        
        merge_data = pd.concat(data_all)
        merge_outline = pd.concat(outline_all)

        # 画像入りデータの保存（BOX）
        make_output_data(merge_data,download_path,save_name=sv_name) 

        # 画像入りデータにoutline追加
        path_to_excel = f'{download_path}/{sv_name}.xlsx'  
        add_df_to_excel(path_to_excel,add_dataframe = merge_outline,sheet_n = 'outline')

        # outlinenの表示
        st.dataframe(merge_outline)

        # 画像入りデータフレームの表示
        st.data_editor(merge_data,column_config={"apps": st.column_config.ImageColumn("Preview Image", help="Streamlit app preview screenshots")},hide_index=True,)

        # ダウンロード
        #　画像あり（BOX経由）
        st.markdown(""" #### 下記リンクをコピーしブラウザに貼付けてください """)
        st.code(f"file:///D:/Box/[Azure_box]_MI-ATRL-3/bio_lab/Antismash/{sv_name}.xlsx")
        # 画像なし
        merge_data.to_excel(buf := BytesIO(), index=True)
        st.download_button("Download (画像なし)",buf.getvalue(),f"{sv_name}.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    

st.title('Predicting BioActive agent tool')
st.image(f'{i_path}/image.png')

st.sidebar.subheader('STEP1: ゲノム情報の取得')
select = st.sidebar.selectbox('入手先',['-select-','NCBI'])

st.sidebar.subheader('STEP2: 二次代謝遺伝子の予測')
select1 = st.sidebar.selectbox('選択',['-select-','Anti_smash','Predict_function'])

st.sidebar.subheader('STEP3: Docking simulation')
select2 = st.sidebar.selectbox('選択',['-select-','autodock_vina'])

st.sidebar.subheader('STEP4: データ統合')
select3 = st.sidebar.selectbox('選択',['-select-','antismash','antismash/pred_func','antismash/pred_func/docking'])

if select == 'NCBI':
   gain_genome()

if select1 == 'Anti_smash':
   anti_smash()

if select1 == 'Predict_function':
   add_research_article()

if select2 == 'autodock_vina':
   docking()

if select3 == 'antismash':
   create_final_data('Antismash')

if select3 == 'antismash/pred_func':
   create_final_data('Pred_func')